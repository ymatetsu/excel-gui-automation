import openpyxl

def extract_table_from_excel(file_path, id_text, start_text):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Sheet1"]

    id_cell = None
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value == id_text:
                id_cell = cell
                break
        if id_cell:
            break
    if not id_cell:
        raise ValueError(f"対象セル '{id_text}' が見つかりませんでした")

    start_cell = None
    for row in range(id_cell.row + 1, ws.max_row + 1):
        value = ws.cell(row=row, column=id_cell.column).value
        if value == start_text:
            start_cell = ws.cell(row=row, column=id_cell.column)
            break
    if not start_cell:
        raise ValueError(f"'{start_text}' セルが見つかりませんでした")

    start_row, start_col = start_cell.row, start_cell.column

    end_col = start_col
    for c in range(start_col, ws.max_column + 1):
        if ws.cell(start_row, c).value is None:
            break
        end_col = c

    end_row = start_row
    for r in range(start_row, ws.max_row + 1):
        row_values = [ws.cell(r, c).value for c in range(start_col, end_col + 1)]
        if all(v is None for v in row_values):
            break
        end_row = r

    headers = [ws.cell(start_row, c).value for c in range(start_col, end_col + 1)]

    table_data = []
    for r in range(start_row + 1, end_row + 1):
        row_dict = {}
        for idx, c in enumerate(range(start_col, end_col + 1)):
            row_dict[headers[idx]] = ws.cell(r, c).value
        table_data.append(row_dict)

    return table_data
